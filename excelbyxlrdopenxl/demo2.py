from openpyxl import Workbook, load_workbook

wb = load_workbook('Book1.xlsx')
ws = wb.active
print(ws)
print(ws['A1'])
print(ws['A1'].value)